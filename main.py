# -*- coding: utf-8 -*-
"""
PCAN .trc -> Excel .xlsx (tabulato) con:
- prime 6 righe del file copiate "as-is" in alto;
- colonna "Δt since same ID (ms)" subito dopo "Time Offset (ms)" (diff per ID);
- colonna "Δt ERR (>X% from mean)" con formula Excel che usa soglia in G6 (percentuale);
- evidenziazione righe in base a 3 ID impostabili in G2, G3, G4 (contains);
- apertura automatica del file Excel generato.

Colori:
- G2 + righe corrispondenti: rosso chiaro
- G3 + righe corrispondenti: verde chiaro
- G4 + righe corrispondenti: azzurro
"""

import os
import sys
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


def parse_trc(trc_text: str):
    """Parsa un PCAN-View .trc (v5) in (prime 6 righe info, records tabella)."""
    lines = trc_text.splitlines()
    info_lines = lines[:6]  # mantieni le prime 6 righe così come sono

    records = []
    for ln in lines[6:]:
        if ln.strip().startswith(";"):
            continue  # ignora righe commento
        # Esempio riga messaggio:
        # "     2)         0.8  Rx         0086  8  00 80 15 00 00 00 00 00"
        if ")" in ln:
            try:
                clean = ln.strip().replace(")", "")
                parts = clean.split()
                # parts atteso: [msg_no, time_offset, type, id_hex, dlc, byte0..byte7]
                if len(parts) >= 5 and parts[0].isdigit():
                    msg_no = int(parts[0])
                    time_offset = float(parts[1].replace(",", "."))  # gestisci virgola decimale
                    msg_type = parts[2]
                    id_hex = parts[3].upper()
                    dlc = int(parts[4])
                    data_bytes = [b.upper() for b in parts[5:5 + 8]]
                    data_bytes += [""] * (8 - len(data_bytes))  # pad a 8 byte

                    records.append({
                        "Message #": msg_no,
                        "Time Offset (ms)": time_offset,
                        "Type": msg_type,
                        "ID (hex)": id_hex,
                        "DLC": dlc,
                        "Byte0": data_bytes[0],
                        "Byte1": data_bytes[1],
                        "Byte2": data_bytes[2],
                        "Byte3": data_bytes[3],
                        "Byte4": data_bytes[4],
                        "Byte5": data_bytes[5],
                        "Byte6": data_bytes[6],
                        "Byte7": data_bytes[7],
                    })
            except Exception:
                # ignora righe malformate
                pass

    records.sort(key=lambda r: r["Message #"])
    return info_lines, records


def add_id_highlights(ws, first_data_row: int, last_data_row: int, id_col_letter: str, last_col_letter: str):
    """Aggiunge 3 regole di formattazione condizionale basate su G2, G3, G4 e colora le celle input."""
    # Etichette di aiuto
    ws["F2"] = "ID filtro #1"
    ws["F3"] = "ID filtro #2"
    ws["F4"] = "ID filtro #3"

    # Riempimenti (ARGB a 8 cifre, con alpha "FF")
    red_fill = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
    green_fill = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")
    blue_fill = PatternFill(start_color="FFCCFFFF", end_color="FFCCFFFF", fill_type="solid")

    # Colora le celle input e impostale come testo per preservare 0 iniziali
    for cell, fill in (("G2", red_fill), ("G3", green_fill), ("G4", blue_fill)):
        ws[cell].fill = fill
        ws[cell].number_format = "@"  # formato testo

    # Soglia ERR in G6 (percentuale)
    ws["F6"] = "ERR soglia"
    ws["G6"] = 0.20
    ws["G6"].number_format = "0%"

    # Intervallo dati dall'inizio all'ultima colonna reale
    data_range = f"A{first_data_row}:{last_col_letter}{last_data_row}"

    # Formule: riferimento alla colonna ID (lettera dinamica), riga relativa
    f1 = f'=AND($G$2<>"",NOT(ISERROR(SEARCH($G$2,${id_col_letter}{first_data_row}))))'
    f2 = f'=AND($G$3<>"",NOT(ISERROR(SEARCH($G$3,${id_col_letter}{first_data_row}))))'
    f3 = f'=AND($G$4<>"",NOT(ISERROR(SEARCH($G$4,${id_col_letter}{first_data_row}))))'

    ws.conditional_formatting.add(data_range, FormulaRule(formula=[f1], fill=red_fill))
    ws.conditional_formatting.add(data_range, FormulaRule(formula=[f2], fill=green_fill))
    ws.conditional_formatting.add(data_range, FormulaRule(formula=[f3], fill=blue_fill))

    # Blocca riga intestazione per lo scroll
    ws.freeze_panes = f"A{first_data_row}"


def export_xlsx(in_path: Path) -> Path:
    """Legge .trc e scrive .xlsx con tabella, Δt per ID, colonna ERR (formula) e formattazione condizionale."""
    text = in_path.read_text(errors="replace")
    info_lines, records = parse_trc(text)

    # DataFrame base
    df = pd.DataFrame(records, columns=[
        "Message #", "Time Offset (ms)", "Type", "ID (hex)", "DLC",
        "Byte0", "Byte1", "Byte2", "Byte3", "Byte4", "Byte5", "Byte6", "Byte7"
    ])

    # Δt per stesso ID (stessa unit del time offset). Prima occorrenza = NaN.
    dtime = df.groupby("ID (hex)")["Time Offset (ms)"].diff()

    # Inserisci colonne: Δt e, SUBITO DOPO, la colonna ERR (che poi riempiremo con formula)
    df.insert(2, "Δt since same ID (ms)", dtime)
    df.insert(3, "Δt ERR (>X% from mean)", "")  # placeholder; formule Excel verranno scritte dopo

    out_path = in_path.with_suffix(".xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        start_row = 7  # -> header a riga 8, dati da riga 9
        df.to_excel(writer, sheet_name="Trace", index=False, startrow=start_row)
        ws = writer.sheets["Trace"]

        # Scrivi le prime 6 righe "as-is"
        for i, line in enumerate(info_lines, start=1):
            ws.cell(row=i, column=1, value=line)

        # Mappa header -> lettera colonna (dopo l'inserimento)
        headers = list(df.columns)
        # posizioni (1-based) utili
        id_idx = headers.index("ID (hex)") + 1
        dt_idx = headers.index("Δt since same ID (ms)") + 1
        err_idx = headers.index("Δt ERR (>X% from mean)") + 1
        last_idx = len(headers)

        id_col_letter = get_column_letter(id_idx)
        dt_col_letter = get_column_letter(dt_idx)
        err_col_letter = get_column_letter(err_idx)
        last_col_letter = get_column_letter(last_idx)

        # Larghezze colonne sensate
        desired_widths = {
            "Message #": 12,
            "Time Offset (ms)": 16,
            "Δt since same ID (ms)": 18,
            "Δt ERR (>X% from mean)": 20,
            "Type": 10,
            "ID (hex)": 12,
            "DLC": 8,
            "Byte0": 8, "Byte1": 8, "Byte2": 8, "Byte3": 8,
            "Byte4": 8, "Byte5": 8, "Byte6": 8, "Byte7": 8,
        }
        for idx, name in enumerate(headers, start=1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = desired_widths.get(name, 12)

        # AutoFilter sulla riga header (copre tutte le colonne)
        ws.auto_filter.ref = f"A{start_row+1}:{last_col_letter}{start_row+1}"

        # Range righe dati
        first_data_row = start_row + 2      # 9
        last_data_row = first_data_row + len(df) - 1 if len(df) else first_data_row

        # Formattazione condizionale + input helper (G2..G4 + soglia G6)
        add_id_highlights(ws, first_data_row, last_data_row, id_col_letter, last_col_letter)

        # ---- Formula ERR per ogni riga dati ----
        # Media Δt per lo stesso ID via AVERAGEIFS su tutta la colonna Δt (righe dati).
        # mean_expr = IFERROR(AVERAGEIFS(dt_range, id_range, id_cell), 0)
        dt_range = f"${dt_col_letter}${first_data_row}:${dt_col_letter}${last_data_row}"
        id_range = f"${id_col_letter}${first_data_row}:${id_col_letter}${last_data_row}"

        for r in range(first_data_row, last_data_row + 1):
            dt_cell = f"${dt_col_letter}{r}"
            id_cell = f"${id_col_letter}{r}"
            mean_expr = f"IFERROR(AVERAGEIFS({dt_range},{id_range},{id_cell}),0)"
            # IF( OR(dt="", mean<=0), "", IF( ABS(dt-mean) > ($G$6)*mean, "ERR","") )
            formula = (
                f'=IF(OR({dt_cell}="",{mean_expr}<=0),"",'
                f'IF(ABS({dt_cell}-{mean_expr})>($G$6)*{mean_expr},"ERR",""))'
            )
            ws[f"{err_col_letter}{r}"].value = formula

    return out_path


def open_file_with_default_app(path: Path):
    """Apre il file con l'applicazione di default del sistema operativo."""
    try:
        if sys.platform.startswith("win"):
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])
    except Exception as e:
        messagebox.showwarning("Apertura file",
                               f"File creato ma non sono riuscito ad aprirlo automaticamente:\n{e}")


def main():
    root = tk.Tk()
    root.withdraw()  # niente finestra principale

    path = filedialog.askopenfilename(
        title="Seleziona il file PCAN Trace (.trc)",
        filetypes=[("PCAN Trace", "*.trc"), ("Testo", "*.txt"), ("Tutti i file", "*.*")]
    )
    if not path:
        return

    try:
        out = export_xlsx(Path(path))
        messagebox.showinfo(
            "Operazione completata",
            f"Excel creato:\n{out}\n\n"
            "Inserisci gli ID in G2, G3, G4 per evidenziare le righe. "
            "Modifica la soglia in G6 (es. 10% / 30%)."
        )
        open_file_with_default_app(Path(out))
    except Exception as e:
        messagebox.showerror("Errore", f"Impossibile creare l'Excel:\n{e}")


if __name__ == "__main__":
    main()
